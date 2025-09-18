import pandas as pd # type: ignore
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import Alignment, PatternFill # type: ignore

# Read files without considering the first row as column headers
file1 = pd.read_excel('file1.xlsx', header=None)
file2 = pd.read_excel('file2.xlsx', header=None)

# Use the first row as column headers
headers_1 = file1.iloc[0].astype(str).str.strip().tolist()
headers_2 = file2.iloc[0].astype(str).str.strip().tolist()

# Data from the second row onward (without header row)
data1 = file1.iloc[1:].reset_index(drop=True)
data2 = file2.iloc[1:].reset_index(drop=True)

# Set column names for data
data1.columns = headers_1
data2.columns = headers_2

# Sort first by first column, then eighth, then seventh
sort_cols_1 = [headers_1[0], headers_1[7], headers_1[6]]
sort_cols_2 = [headers_2[0], headers_2[7], headers_2[6]]

data1 = data1.sort_values(by=sort_cols_1).reset_index(drop=True)
data2 = data2.sort_values(by=sort_cols_2).reset_index(drop=True)

# Key column name and third column name for comparison
key_column_name_1 = headers_1[0]
key_column_name_2 = headers_2[0]
compare_column_name_1 = headers_1[2]
compare_column_name_2 = headers_2[2]
compare_column_name_7_1 = headers_1[6]  # Column 7 (index 6)
compare_column_name_7_2 = headers_2[6]  # Column 7 (index 6)

# Ensure key and comparison column values are strings
data1[key_column_name_1] = data1[key_column_name_1].astype(str).str.strip()
data2[key_column_name_2] = data2[key_column_name_2].astype(str).str.strip()
data1[compare_column_name_1] = data1[compare_column_name_1].astype(str).str.strip()
data2[compare_column_name_2] = data2[compare_column_name_2].astype(str).str.strip()
data1[compare_column_name_7_1] = data1[compare_column_name_7_1].astype(str).str.strip()
data2[compare_column_name_7_2] = data2[compare_column_name_7_2].astype(str).str.strip()

# Find common keys based on first column, third column, and seventh column
common_keys = set(
    data1.apply(lambda x: (x[key_column_name_1], x[compare_column_name_1], x[compare_column_name_7_1]), axis=1)
).intersection(
    set(data2.apply(lambda x: (x[key_column_name_2], x[compare_column_name_2], x[compare_column_name_7_2]), axis=1))
)

# Rows only in file 1 and 2
diff_rows_1 = data1[~data1.apply(lambda x: (x[key_column_name_1], x[compare_column_name_1], x[compare_column_name_7_1]), axis=1).isin(
    data2.apply(lambda x: (x[key_column_name_2], x[compare_column_name_2], x[compare_column_name_7_2]), axis=1))]

diff_rows_2 = data2[~data2.apply(lambda x: (x[key_column_name_2], x[compare_column_name_2], x[compare_column_name_7_2]), axis=1).isin(
    data1.apply(lambda x: (x[key_column_name_1], x[compare_column_name_1], x[compare_column_name_7_1]), axis=1))]

# Common rows
common_rows_1 = data1[data1.apply(lambda x: (x[key_column_name_1], x[compare_column_name_1], x[compare_column_name_7_1]), axis=1).isin(
    data2.apply(lambda x: (x[key_column_name_2], x[compare_column_name_2], x[compare_column_name_7_2]), axis=1))]

common_rows_2 = data2[data2.apply(lambda x: (x[key_column_name_2], x[compare_column_name_2], x[compare_column_name_7_2]), axis=1).isin(
    data1.apply(lambda x: (x[key_column_name_1], x[compare_column_name_1], x[compare_column_name_7_1]), axis=1))]

# Save to Excel
output_file = 'comparison_result.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    common_rows_1.to_excel(writer, sheet_name='file1 - common', index=False)
    common_rows_2.to_excel(writer, sheet_name='file2 - common', index=False)
    diff_rows_1.to_excel(writer, sheet_name='Just in file1', index=False)
    diff_rows_2.to_excel(writer, sheet_name='Just in file2', index=False)
    
# Open file for right alignment, RTL, and coloring
wb = load_workbook(output_file)
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Right-align content
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='right')
    # Set sheet to right-to-left
    ws.sheet_view.rightToLeft = True

    if 'common' in sheet_name:
        first_green_done = False  # Only first row green
        for i, row in enumerate(ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column), start=2):
            cell_h = ws.cell(row=i, column=8)  # Column H
            if cell_h.value is None or str(cell_h.value).strip() == '':
                # Column 8 empty → Color entire row orange
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=i, column=col).fill = orange_fill

wb.save(output_file)
print("file created as: comparison_result.xlsx")